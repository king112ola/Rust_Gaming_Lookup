from os import write
from bs4.element import CharsetMetaAttributeValue
import requests as rq
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import numpy as np



url = ""
urlss = ''
count = 0
countForList = 0

with open('config.txt', 'r') as reader:
    excelStartLocation = reader.readline()
    excelStartLocation = int(excelStartLocation)
    urlss2 = list(reader.read().split("\n"))
    urlss = [x.strip(' ') for x in urlss2]
    reader.close()

with open("config.txt", "w") as f:
    f.writelines(str(excelStartLocation+(urlss.__len__()))+'\n')
    f.write('\n'.join(urlss))
    f.close()
    print(urlss.__len__())
    

w, h = 5, (urlss.__len__()+1);
arr = [[0 for x in range(w)] for y in range(h)]     

wb = load_workbook('RustFrameworks.xlsx')
ws = wb.active



for i in range(urlss.__len__()):
    

    url = urlss[countForList]
    countForList += 1

    
    response = rq.get(url)  
    html_doc = response.text 
    soup = BeautifulSoup(response.text, "lxml")
    print()
    
    desp = ""
     
    try:
        desp = ((soup.find("p",class_='f4 mt-3').getText().strip()))
    except:
        desp = ""
   
    Name= ""

    try:
        Name = (soup.find("strong", class_= "mr-2 flex-self-stretch").getText().strip())
    except:
        Name = ""
    
    
    checkForAdapt =  desp
    adapt = ""
    types = ""
    
    print("Name: "+Name )
    print("URL: "+url)
    print("Desc: "+desp)
    
    if "bevy" in (checkForAdapt.lower() or Name.lower()): 
        adapt = "BEVY"
        print("Adaptive Engine: " +adapt)
    elif "amethyst" in (checkForAdapt.lower() or Name.lower()):
        adapt = "Amethyst"
        print("Adaptive Engine: " +adapt)
    else:
        print("Adaptive Engine: ")

    if ("2d" or "2-dime" or "two-dime") in (checkForAdapt.lower()) and ("3d" or "3-dime" or "three-dime") in (checkForAdapt.lower()): 
        print("Types: 2D / 3D")
        types = "2D / 3D"
    elif ("2d" or "2-dime" or "two-dime") in (checkForAdapt.lower() ): 
        print("Types: 2D")
        types = "2D"
    elif ("3d" or "3-dime" or "three-dime") in (checkForAdapt.lower() ): 
        print("Types: 3D")
        types = "3D"
    else:
        print("Types: NULL\n")
        types = ""

    for j in range(5):
        
        if j==0:
            try:
                arr[count][j] = Name
            except:
                arr[count][j] = np.nan
        elif j==1:
            try:
                arr[count][j] = url
            except:
                arr[count][j] = np.nan
        elif j==2:
            try:
                arr[count][j] = desp
            except:
                arr[count][j] = np.nan
        elif j==3:
            try:
                arr[count][j] = adapt
            except:
                arr[count][j] = np.nan
        elif j==4:
            try:
                arr[count][j] = types
            except:
                arr[count][j] = np.nan

    for col in range(1,6):
            char = get_column_letter(col+9)
            if "/github.com/" in str(arr[count][col-1]):
                 ws[char+str(excelStartLocation)].hyperlink = str(arr[count][col-1])
            else:
                ws[char+str(excelStartLocation)].value = str(arr[count][col-1])
 
    count += 1
    excelStartLocation += 1

wb.save('RustFrameworks.xlsx')


